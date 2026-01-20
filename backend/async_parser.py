#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sqlite3
import os
import re
import json
from datetime import datetime
import sys

# 获取当前目录（支持打包后运行）
if getattr(sys, 'frozen', False):
    # 打包后运行
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 开发模式运行
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 数据库配置
DATABASE = os.path.join(BASE_DIR, 'api_generator.db')

# 异步文件解析函数
def parse_file_async(file_id, file_path, file_name, file_content_type):
    """异步解析文件，提取接口信息"""
    conn = None
    parsed_interfaces = 0
    parsed_params = 0
    parsed_responses = 0
    content = ""
    
    try:
        # 获取文件扩展名
        file_ext = os.path.splitext(file_name)[1].lower()
        supported_text_types = ['.json', '.md', '.txt']
        supported_binary_types = ['.docx', '.xlsx', '.pdf', '.png', '.jpg', '.jpeg']
        
        # 根据文件类型读取内容
        if file_ext in supported_text_types:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        elif file_ext in supported_binary_types:
            # 二进制文件，先保存，后续根据类型解析
            if file_ext == '.docx':
                try:
                    # 解析DOCX文件
                    from docx import Document
                    doc = Document(file_path)
                    for para in doc.paragraphs:
                        content += para.text + '\n'
                except ImportError:
                    print("警告：未安装python-docx库，无法解析DOCX文件")
                    content = ""
            elif file_ext == '.xlsx':
                try:
                    # 解析XLSX文件
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        # 将每行数据转换为字符串，用制表符分隔
                        row_str = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                        content += row_str + '\n'
                except ImportError:
                    print("警告：未安装openpyxl库，无法解析XLSX文件")
                    content = ""
            elif file_ext == '.pdf':
                try:
                    # 解析PDF文件
                    from PyPDF2 import PdfReader
                    reader = PdfReader(file_path)
                    for page in reader.pages:
                        content += page.extract_text() + '\n'
                except ImportError:
                    print("警告：未安装PyPDF2库，无法解析PDF文件")
                    content = ""
            elif file_ext in ['.png', '.jpg', '.jpeg']:
                try:
                    # 解析图片文件，使用OCR
                    from PIL import Image
                    import pytesseract
                    
                    # 调整图片大小，提高OCR速度
                    img = Image.open(file_path)
                    if img.size[0] > 1920 or img.size[1] > 1080:
                        img.thumbnail((1920, 1080))
                    
                    # 增强图片预处理，提高OCR准确性
                    # 1. 转换为灰度图
                    img = img.convert('L')
                    
                    # 2. 调整对比度
                    from PIL import ImageEnhance
                    enhancer = ImageEnhance.Contrast(img)
                    img = enhancer.enhance(1.5)
                    
                    # 3. 调整亮度
                    enhancer = ImageEnhance.Brightness(img)
                    img = enhancer.enhance(1.2)
                    
                    # 4. 锐化图片
                    enhancer = ImageEnhance.Sharpness(img)
                    img = enhancer.enhance(1.5)
                    
                    # 使用更优的OCR参数
                    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_./\s\n:#*+(){}[]|"\'\`'
                    content = pytesseract.image_to_string(img, lang='chi_sim+eng', config=custom_config)
                    
                    # 添加调试信息，输出OCR识别的文本内容
                    print(f"\n=== 图片OCR识别结果 ===")
                    print(f"图片文件: {file_name}")
                    print(f"OCR识别的文本长度: {len(content)} 字符")
                    print(f"OCR识别的文本内容:\n{content[:2000]}...")
                    print("==================\n")
                    
                except ImportError as e:
                    print(f"警告：无法解析图片文件，可能需要安装依赖: {e}")
                    content = ""
                except Exception as e:
                    print(f"解析图片文件失败: {e}")
                    import traceback
                    traceback.print_exc()
                    content = ""
            else:
                # 其他二进制文件，标记为未解析
                content = ""
        else:
            # 不支持的文件类型，标记为未解析
            content = ""
        
        # 解析接口信息
        # 首先将文件内容按行读取，然后寻找接口定义
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # 尝试直接解析JSON内容
        interface_matches = []
        swagger_data = None
        
        # 尝试直接解析JSON内容
        try:
            # 尝试解析JSON文件
            json_data = json.loads(content)
            
            # 尝试解析Swagger/OpenAPI规范
            if 'paths' in json_data:
                print("检测到Swagger/OpenAPI规范的JSON文件")
                swagger_data = json_data
                
                # 遍历所有路径
                for path, methods in json_data['paths'].items():
                    # 遍历该路径下的所有请求方法
                    for method, details in methods.items():
                        # 只处理HTTP方法
                        if method in ['get', 'post', 'put', 'delete', 'patch']:
                            method = method.upper()
                            # 提取接口名称
                            name = details.get('summary', f'{method} {path}')
                            # 添加到匹配结果中，使用特殊的标记表示这是从JSON解析的
                            interface_matches.append((None, name, method, path, swagger_data, details))
            else:
                print("JSON文件格式不支持，尝试Markdown解析")
                # 即使是JSON文件，也继续执行Markdown解析的逻辑，确保能提取到接口
        except json.JSONDecodeError:
            print("JSON解析失败，尝试Markdown解析")
            # JSON解析失败，继续使用Markdown解析
            pass
        
        # 如果没有从JSON解析到接口，尝试使用Markdown解析
        if not interface_matches:
            # 使用更可靠的正则表达式直接匹配所有接口定义
            # 匹配格式：### 接口名称
            # **POST** /path 或 ### 接口名称
            # POST /path
            interface_pattern = r'^(#{2,3})\s+(.+?)\s*\n\s*(?:\*\*)?\s*(GET|POST|PUT|DELETE|PATCH)\s*(?:\*\*)?\s*[`]?([^\s`]+)[`]?' 
            markdown_matches = re.findall(interface_pattern, content, re.MULTILINE)
            # 转换为与JSON解析相同的格式
            interface_matches = [(header, name, method, path, None, None) for header, name, method, path in markdown_matches]
        
        print(f"Markdown格式匹配到 {len(interface_matches)} 个接口")
        
        # 尝试匹配图片中的接口定义格式
        if file_ext in ['.png', '.jpg', '.jpeg']:
            print("\n=== 图片接口匹配开始 ===")
            
            # 图片专用的接口匹配模式列表
            image_patterns = [
                # 1. Markdown格式匹配（### 接口名称 **POST** /path）
                r'^(#{2,3})\s+(.+?)\s*(?:\n|$)[\s\S]*?(?:\*\*)?\s*(GET|POST|PUT|DELETE|PATCH)\s*(?:\*\*)?\s*[`‘’]?([^\s`‘’]+)[`‘’]?',
                # 2. 前端接口映射格式（cashOuterProductProductManagerQuery; /cash/outer/product/productManagerQuery;/, // 查询产品投资经理列表）
                r'(\w+)\s*;\s*([^\s,]+)\s*;\s*([^\s,]+)\s*,\s*//\s*(.+)',
                # 3. 通用HTTP方法+路径匹配（GET /api/v1/users）
                r'\s*(GET|POST|PUT|DELETE|PATCH)\s+([^\s\n]+)\s*',
                # 4. 带描述的HTTP方法+路径匹配（POST /api/v1/login - 用户登录）
                r'\s*(GET|POST|PUT|DELETE|PATCH)\s+([^\s\n]+)\s*[-\s]+(.+?)\s*(?:\n|$)',
                # 5. 代码块中的接口定义（app.post('/api/v1/users', ...)）
                r"\w+\s*\.(get|post|put|delete|patch)\s*\(['\"]([^'\"]+)['\"]",
                # 6. 表格格式的接口定义
                r'\|\s*(GET|POST|PUT|DELETE|PATCH)\s*\|\s*([^\|]+)\s*\|\s*([^\|]+)\s*\|',
                # 7. 简化的接口定义（POST /api/v1/users）
                r'^(?:\s*|\d+\.)\s*(GET|POST|PUT|DELETE|PATCH)\s+([^\s\n]+)\s*$'
            ]
            
            # 记录已经匹配到的接口，避免重复
            matched_interfaces = set()
            
            # 遍历所有匹配模式
            for pattern_index, pattern in enumerate(image_patterns):
                try:
                    pattern_matches = re.findall(pattern, content, re.MULTILINE | re.DOTALL)
                    print(f"图片模式 {pattern_index+1} 匹配到 {len(pattern_matches)} 个结果")
                    
                    for match in pattern_matches:
                        try:
                            # 根据不同模式解析匹配结果
                            if pattern_index == 0:  # Markdown格式
                                header_level_str, name, method, path = match
                                description = ""
                            elif pattern_index == 1:  # 前端接口映射格式
                                name, path, _, description = match
                                # 从名称或描述中推断方法
                                method = 'GET'
                                if 'post' in name.lower() or 'post' in description.lower():
                                    method = 'POST'
                                elif 'put' in name.lower() or 'put' in description.lower():
                                    method = 'PUT'
                                elif 'delete' in name.lower() or 'delete' in description.lower():
                                    method = 'DELETE'
                                elif 'patch' in name.lower() or 'patch' in description.lower():
                                    method = 'PATCH'
                            elif pattern_index == 2 or pattern_index == 6:  # 通用HTTP方法+路径匹配
                                method, path = match
                                name = f"{method} {path}"
                                description = ""
                            elif pattern_index == 3:  # 带描述的HTTP方法+路径匹配
                                method, path, description = match
                                name = f"{method} {path}"
                            elif pattern_index == 4:  # 代码块中的接口定义
                                method, path = match
                                method = method.upper()
                                name = f"{method} {path}"
                                description = ""
                            elif pattern_index == 5:  # 表格格式
                                method, path, description = match
                                name = f"{method} {path}"
                            else:
                                continue
                            
                            # 清理数据
                            name = name.strip()
                            method = method.upper()
                            path = path.strip()
                            description = description.strip()
                            
                            # 跳过无效的路径
                            if not path or len(path) < 2:
                                continue
                            
                            # 如果路径是完整URL，只保留路径部分
                            if path.startswith('http'):
                                from urllib.parse import urlparse
                                parsed_url = urlparse(path)
                                path = parsed_url.path
                            
                            # 生成唯一标识符，避免重复
                            interface_key = f"{method} {path}"
                            if interface_key in matched_interfaces:
                                continue
                            matched_interfaces.add(interface_key)
                            
                            # 打印调试信息
                            print(f"从图片中解析到接口: {name} {method} {path}")
                            
                            # 添加到匹配结果中
                            interface_matches.append((None, name, method, path, None, None))
                        except Exception as e:
                            print(f"解析匹配结果失败: {e}")
                            import traceback
                            traceback.print_exc()
                except Exception as e:
                    print(f"执行匹配模式 {pattern_index+1} 失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            print(f"图片解析共匹配到 {len(matched_interfaces)} 个唯一接口")
            print("=== 图片接口匹配结束 ===")
        
        for i, match in enumerate(interface_matches):
            header_level_str, name, method, path, swagger_data, swagger_details = match
            
            # 清理数据
            name = name.strip()
            method = method.upper()
            path = path.strip()
            
            # 处理从JSON解析的接口（header_level_str为None）
            header_level = 3  # 默认值为3
            if header_level_str is not None:
                # 将header_level_str转换为整数（例如 '##' -> 2）
                header_level = len(header_level_str)
            
            # 如果路径是完整URL，只保留路径部分
            if path.startswith('http'):
                from urllib.parse import urlparse
                parsed_url = urlparse(path)
                path = parsed_url.path
            
            # 打印调试信息
            print(f"解析到接口: {name} {method} {path}, 标题级别: {header_level}")
            
            # 创建接口记录
            cursor.execute('''
                INSERT INTO interfaces (name, path, method, description, file_id)
                VALUES (?, ?, ?, ?, ?)
            ''', (name, path, method, '', file_id))
            interface_id = cursor.lastrowid
            parsed_interfaces += 1
            
            # 创建默认Mock配置
            cursor.execute('''
                INSERT INTO mock_configs (interface_id, enabled, default_count)
                VALUES (?, ?, ?)
            ''', (interface_id, 1, 10))
            
            # 初始化请求参数和响应参数列表
            request_params = []
            response_fields = []
            
            # 处理Markdown文件
            if header_level_str is not None:
                # 匹配当前接口的起始位置
                # 先找到接口标题的位置
                title_pattern = r'^' + re.escape(f"{'#' * header_level} {name}") + r'\s*$'
                title_match = re.search(title_pattern, content, re.MULTILINE)
                if title_match:
                    # 从当前接口的标题位置开始提取内容
                    interface_start = title_match.start()
                    # 获取当前标题的结束位置
                    interface_title_end = title_match.end()
                    
                    # 找到下一个接口的起始位置，从当前标题结束位置开始搜索
                    next_interface_start = len(content)
                    next_interface_pattern = r'^#{2,}\s+'  # 匹配下一个接口标题
                    next_matches = list(re.finditer(next_interface_pattern, content[interface_title_end:], re.MULTILINE))
                    if next_matches:
                        next_interface_start = interface_title_end + next_matches[0].start()
                    
                    # 提取当前接口的完整内容
                    interface_content = content[interface_start:next_interface_start]
                    
                    # 提取请求参数，支持多种格式
                    print(f"=== 提取Markdown请求参数 ===")
                    # 尝试多种参数格式
                    param_patterns = [
                        r'\*\*参数：\*\*\s*```js\s*([\s\S]*?)```',  # JavaScript代码块格式
                        r'\*\*参数：\*\*\s*```json\s*([\s\S]*?)```',  # JSON代码块格式
                        r'\*\*参数：\*\*\s*```\s*([\s\S]*?)```',  # 无语言标识的代码块格式
                    ]
                    
                    param_code = None
                    for pattern in param_patterns:
                        param_match = re.search(pattern, interface_content, re.DOTALL)
                        if param_match:
                            param_code = param_match.group(1)
                            print(f"匹配到参数格式，提取到参数代码块")
                            break
                    
                    if param_code:
                        # 移除可能的变量声明和结束符
                        param_code_clean = param_code.replace('let params =', '').replace('const params =', '').replace('var params =', '').rstrip(';').strip()
                        
                        try:
                            # 解析JSON
                            param_json = json.loads(param_code_clean)
                            print(f"成功解析JSON参数")
                            # 提取参数
                            for param_name, param_value in param_json.items():
                                # 简单参数，提取类型
                                param_type = 'string'  # 默认类型
                                if isinstance(param_value, str):
                                    param_type = 'string'
                                elif isinstance(param_value, int):
                                    param_type = 'int'
                                elif isinstance(param_value, bool):
                                    param_type = 'boolean'
                                elif isinstance(param_value, float):
                                    param_type = 'double'
                                elif isinstance(param_value, list):
                                    param_type = 'list'
                                elif isinstance(param_value, dict):
                                    param_type = 'object'
                                
                                # 提取注释中的描述
                                description = ''
                                comment_pattern = f"{param_name}:[^/]*//(.*)"
                                comment_match = re.search(comment_pattern, param_code)
                                if comment_match:
                                    description = comment_match.group(1).strip()
                                
                                # 提取示例值
                                example = str(param_value)
                                if isinstance(param_value, list):
                                    example = '[]'
                                elif isinstance(param_value, dict):
                                    example = '{}'
                                
                                request_params.append((param_name, param_type, 1, description, example))
                                parsed_params += 1
                                print(f"提取到请求参数: {param_name} ({param_type})")
                        except json.JSONDecodeError as e:
                            print(f"JSON解析失败，尝试手动解析: {e}")
                            # 手动解析优化，支持更复杂的格式
                            # 1. 移除首尾的大括号和空格
                            if param_code_clean.startswith('{'):
                                param_code_clean = param_code_clean[1:].strip()
                            if param_code_clean.endswith('}'):
                                param_code_clean = param_code_clean[:-1].strip()
                            
                            # 2. 按行分割
                            lines = param_code_clean.split('\n')
                            for line in lines:
                                line = line.strip()
                                if not line or line.startswith('//'):
                                    continue
                                
                                # 3. 分割键值对，处理可能的逗号
                                if ':' in line:
                                    # 分割键和值，只分割第一个冒号
                                    key_part, value_part = line.split(':', 1)
                                    param_name = key_part.strip()
                                    
                                    # 处理值部分，移除逗号和注释
                                    value_part = value_part.split('//')[0].strip().rstrip(',')
                                    
                                    # 提取描述（如果有）
                                    description = ''
                                    comment_match = re.search(r'//(.*)', line)
                                    if comment_match:
                                        description = comment_match.group(1).strip()
                                    
                                    # 确定参数类型和示例值
                                    param_type = 'string'
                                    example = ''
                                    
                                    # 处理空值
                                    if value_part in ['', "''", '""']:
                                        param_type = 'string'
                                        example = ''
                                    # 处理布尔值
                                    elif value_part.lower() in ['true', 'false']:
                                        param_type = 'boolean'
                                        example = value_part.lower()
                                    # 处理数字
                                    elif re.match(r'^\d+$', value_part):
                                        param_type = 'int'
                                        example = value_part
                                    elif re.match(r'^\d+\.\d+$', value_part):
                                        param_type = 'double'
                                        example = value_part
                                    # 处理列表
                                    elif value_part.startswith('[') and value_part.endswith(']'):
                                        param_type = 'list'
                                        example = '[]'
                                    # 处理对象
                                    elif value_part.startswith('{') and value_part.endswith('}'):
                                        param_type = 'object'
                                        example = '{}'
                                    # 处理字符串
                                    else:
                                        param_type = 'string'
                                        # 移除引号
                                        if (value_part.startswith("'") and value_part.endswith("'") or 
                                           (value_part.startswith('"') and value_part.endswith('"'))):
                                            example = value_part[1:-1]
                                        else:
                                            example = value_part
                                    
                                    request_params.append((param_name, param_type, 1, description, example))
                                    parsed_params += 1
                                    print(f"手动提取到请求参数: {param_name} ({param_type})")
                    else:
                        # 尝试匹配表格格式的参数
                        table_param_pattern = r'\*\*参数：\*\*\s*\n\|(.+?)\|(.+?)\|(.+?)\|(.+?)\|(.+?)\|'
                        table_param_match = re.search(table_param_pattern, interface_content, re.DOTALL)
                        if table_param_match:
                            print(f"匹配到表格格式参数")
                            # 处理表格格式参数（此处简化处理，实际需要更复杂的表格解析）
                            pass
                    
                    # 提取响应参数，支持多种格式
                    print(f"=== 提取Markdown响应参数 ===")
                    # 尝试多种响应体说明格式
                    response_patterns = [
                        r'\*\*响应体说明：\*\*\s*([\s\S]*?)(?:#{2,}|$)',  # 标准格式
                        r'\*\*响应说明：\*\*\s*([\s\S]*?)(?:#{2,}|$)',  # 简化格式
                        r'\*\*响应：\*\*\s*([\s\S]*?)(?:#{2,}|$)',  # 更简化的格式
                    ]
                    
                    response_content = None
                    for pattern in response_patterns:
                        response_match = re.search(pattern, interface_content, re.DOTALL)
                        if response_match:
                            response_content = response_match.group(1)
                            print(f"匹配到响应体说明格式")
                            break
                    
                    if response_content:
                        # 按行解析响应字段
                        lines = response_content.strip().split('\n')
                        for line in lines:
                            line = line.strip()
                            if not line:
                                continue
                            
                            # 使用正则表达式分割，支持多个空格或制表符
                            parts = re.split(r'\s{2,}|\t+', line)
                            if len(parts) >= 3:
                                field_name = parts[0]
                                # 跳过表头行
                                if field_name in ['字段名', '字段', 'name', '参数名', 'key']:
                                    continue
                                
                                # 提取类型（最后一列）
                                field_type = parts[-1]
                                # 提取示例值（第二列）
                                example = parts[1] if len(parts) > 1 else ''
                                # 提取描述（中间列，从第二列到倒数第二列）
                                if len(parts) > 3:
                                    field_description = ' '.join(parts[2:-1])
                                elif len(parts) == 3:
                                    field_description = parts[1] if parts[1] != example else ''
                                else:
                                    field_description = ''
                                
                                # 转换Java类型为通用类型
                                if 'java.lang.String' in field_type or 'java.lang.CharSequence' in field_type or field_type == 'string':
                                    field_type = 'string'
                                elif 'java.lang.Integer' in field_type or 'java.lang.Long' in field_type or 'java.lang.Short' in field_type or 'java.lang.Byte' in field_type or field_type == 'int' or field_type == 'integer':
                                    field_type = 'int'
                                elif 'java.lang.Boolean' in field_type or field_type == 'boolean':
                                    field_type = 'boolean'
                                elif 'java.math.BigDecimal' in field_type or 'java.lang.Double' in field_type or 'java.lang.Float' in field_type or 'java.lang.Number' in field_type or field_type == 'double' or field_type == 'float' or field_type == 'number':
                                    field_type = 'double'
                                elif 'java.lang.Object' in field_type or 'java.util.Map' in field_type or 'java.util.HashMap' in field_type or field_type == 'object':
                                    field_type = 'object'
                                elif 'java.util.List' in field_type or 'java.util.ArrayList' in field_type or 'java.util.Set' in field_type or field_type == 'list' or field_type == 'array':
                                    field_type = 'list'
                                elif 'java.util.Date' in field_type or 'java.time.LocalDate' in field_type or 'java.time.LocalDateTime' in field_type or field_type == 'date' or field_type == 'datetime' or field_type == 'time':
                                    field_type = 'string'
                                else:
                                    field_type = 'string'
                                
                                response_fields.append((field_name, field_type, field_description, example))
                                parsed_responses += 1
                                print(f"提取到响应字段: {field_name} ({field_type})")
            # 处理JSON文件（Swagger/OpenAPI）
            else:
                print(f"处理JSON接口: {name} {method} {path}")
                if swagger_data and swagger_details:
                    # 提取请求参数
                    print(f"=== 提取JSON请求参数 ===")
                    
                    # 1. 处理parameters数组（Swagger 2.0和OpenAPI 3.0兼容）
                    if 'parameters' in swagger_details:
                        for param in swagger_details['parameters']:
                            param_name = param.get('name', '')
                            if not param_name:
                                continue
                            
                            param_type = param.get('type', 'string')
                            required = 1 if param.get('required', False) else 0
                            description = param.get('description', '')
                            example = param.get('example', '')
                            
                            # 处理schema中的类型
                            if 'schema' in param:
                                schema = param['schema']
                                if 'type' in schema:
                                    param_type = schema['type']
                                if 'example' in schema:
                                    example = schema['example']
                            
                            request_params.append((param_name, param_type, required, description, example))
                            parsed_params += 1
                            print(f"提取到请求参数: {param_name} ({param_type})")
                    
                    # 2. 处理requestBody（OpenAPI 3.0+）
                    if 'requestBody' in swagger_details:
                        request_body = swagger_details['requestBody']
                        # 处理content中的媒体类型
                        if 'content' in request_body:
                            for media_type, media_details in request_body['content'].items():
                                # 只处理JSON内容
                                if 'application/json' in media_type:
                                    # 提取schema
                                    if 'schema' in media_details:
                                        schema = media_details['schema']
                                        
                                        # 处理引用
                                        def resolve_ref(ref):
                                            # 简单实现：只处理#/components/schemas/引用
                                            if ref.startswith('#/components/schemas/'):
                                                schema_name = ref.split('/')[-1]
                                                if 'components' in swagger_data and 'schemas' in swagger_data['components']:
                                                    return swagger_data['components']['schemas'].get(schema_name, {})
                                            return {}
                                        
                                        # 递归处理schema中的属性
                                        def process_schema(schema_obj, parent_path=''):
                                            nonlocal parsed_params
                                            # 处理引用
                                            if '$ref' in schema_obj:
                                                resolved_schema = resolve_ref(schema_obj['$ref'])
                                                if resolved_schema:
                                                    return process_schema(resolved_schema, parent_path)
                                                return
                                            
                                            # 处理allOf
                                            if 'allOf' in schema_obj:
                                                for sub_schema in schema_obj['allOf']:
                                                    process_schema(sub_schema, parent_path)
                                                return
                                            
                                            if schema_obj.get('type') == 'object' and 'properties' in schema_obj:
                                                for prop_name, prop_details in schema_obj['properties'].items():
                                                    full_name = f"{parent_path}{prop_name}" if parent_path else prop_name
                                                    prop_type = prop_details.get('type', 'string')
                                                    required = 1 if parent_path or (schema_obj.get('required') and prop_name in schema_obj['required']) else 0
                                                    description = prop_details.get('description', '')
                                                    example = prop_details.get('example', '')
                                                    
                                                    # 递归处理嵌套对象
                                                    if prop_type == 'object' and 'properties' in prop_details:
                                                        process_schema(prop_details, f"{full_name}.")
                                                    elif prop_type == 'array' and 'items' in prop_details:
                                                        items = prop_details['items']
                                                        if items.get('type') == 'object' and 'properties' in items:
                                                            process_schema(items, f"{full_name}[")
                                                        else:
                                                            array_item_type = items.get('type', 'string')
                                                            request_params.append((full_name, f"array[{array_item_type}]", required, description, example))
                                                            parsed_params += 1
                                                            print(f"提取到请求参数: {full_name} (array[{array_item_type}])")
                                                    else:
                                                        request_params.append((full_name, prop_type, required, description, example))
                                                        parsed_params += 1
                                                        print(f"提取到请求参数: {full_name} ({prop_type})")
                                            
                                            elif schema_obj.get('type') == 'array' and 'items' in schema_obj:
                                                items = schema_obj['items']
                                                # 处理数组项
                                                if items.get('type') == 'object' and 'properties' in items:
                                                    process_schema(items, parent_path + '[')
                                                else:
                                                    array_type = items.get('type', 'string')
                                                    required = 1 if parent_path else 0
                                                    description = schema_obj.get('description', '')
                                                    example = schema_obj.get('example', '')
                                                    
                                                    request_params.append((parent_path, f"array[{array_type}]", required, description, example))
                                                    parsed_params += 1
                                                    print(f"提取到请求参数: {parent_path} (array[{array_type}])")
                                        
                                        process_schema(schema)
                    
                    # 提取响应参数
                    print(f"=== 提取JSON响应参数 ===")
                    if 'responses' in swagger_details:
                        # 取第一个成功响应（通常是200）
                        success_response = None
                        for status_code, response_details in swagger_details['responses'].items():
                            if status_code == '200' or status_code.startswith('2'):
                                success_response = response_details
                                break
                        
                        if success_response:
                            # 处理引用
                            def resolve_ref(ref):
                                # 简单实现：只处理#/components/schemas/引用
                                if ref.startswith('#/components/schemas/'):
                                    schema_name = ref.split('/')[-1]
                                    if 'components' in swagger_data and 'schemas' in swagger_data['components']:
                                        return swagger_data['components']['schemas'].get(schema_name, {})
                                return {}
                            
                            # 通用的响应schema处理函数
                            def process_response_schema(schema_obj, parent_path=''):
                                nonlocal parsed_responses
                                # 处理引用
                                if '$ref' in schema_obj:
                                    resolved_schema = resolve_ref(schema_obj['$ref'])
                                    if resolved_schema:
                                        return process_response_schema(resolved_schema, parent_path)
                                    return
                                
                                # 处理allOf
                                if 'allOf' in schema_obj:
                                    for sub_schema in schema_obj['allOf']:
                                        process_response_schema(sub_schema, parent_path)
                                    return
                                
                                if schema_obj.get('type') == 'object' and 'properties' in schema_obj:
                                    for prop_name, prop_details in schema_obj['properties'].items():
                                        full_name = f"{parent_path}{prop_name}" if parent_path else prop_name
                                        prop_type = prop_details.get('type', 'string')
                                        description = prop_details.get('description', '')
                                        example = prop_details.get('example', '')
                                        
                                        # 递归处理嵌套对象
                                        if prop_type == 'object' and 'properties' in prop_details:
                                            process_response_schema(prop_details, f"{full_name}.")
                                        elif prop_type == 'array' and 'items' in prop_details:
                                            items = prop_details['items']
                                            if items.get('type') == 'object' and 'properties' in items:
                                                process_response_schema(items, f"{full_name}[")
                                            else:
                                                array_item_type = items.get('type', 'string')
                                                response_fields.append((full_name, f"array[{array_item_type}]", description, example))
                                                parsed_responses += 1
                                                print(f"提取到响应参数: {full_name} (array[{array_item_type}])")
                                        else:
                                            response_fields.append((full_name, prop_type, description, example))
                                            parsed_responses += 1
                                            print(f"提取到响应参数: {full_name} ({prop_type})")
                                elif schema_obj.get('type') == 'array' and 'items' in schema_obj:
                                    items = schema_obj['items']
                                    if items.get('type') == 'object' and 'properties' in items:
                                        process_response_schema(items, parent_path + '[')
                                    else:
                                        array_type = items.get('type', 'string')
                                        response_fields.append((parent_path, f"array[{array_type}]", schema_obj.get('description', ''), schema_obj.get('example', '')))
                                        parsed_responses += 1
                                        print(f"提取到响应参数: {parent_path} (array[{array_type}])")
                            
                            # 处理Swagger 2.0的响应schema
                            if 'schema' in success_response:
                                schema = success_response['schema']
                                process_response_schema(schema)
                            
                            # 处理OpenAPI 3.0+的响应content
                            elif 'content' in success_response:
                                for media_type, media_details in success_response['content'].items():
                                    if 'application/json' in media_type:
                                        if 'schema' in media_details:
                                            schema = media_details['schema']
                                            process_response_schema(schema)
            # 使用默认的请求参数和响应参数
            pass
            
            # 保存请求参数到数据库
            if len(request_params) > 0:
                for param_name, param_type, required, description, example in request_params:
                    cursor.execute('''
                        INSERT INTO interface_params (name, param_type, required, description, example, interface_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (param_name, param_type, required, description, example, interface_id))
            else:
                # 添加默认参数
                default_params = [
                    ('id', 'string', 1, '唯一标识符', '123456'),
                    ('page', 'int', 0, '页码', '1'),
                    ('page_size', 'int', 0, '每页条数', '10')
                ]
                for param_name, param_type, required, description, example in default_params:
                    cursor.execute('''
                        INSERT INTO interface_params (name, param_type, required, description, example, interface_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (param_name, param_type, required, description, example, interface_id))
            
            # 保存响应参数到数据库
            if len(response_fields) > 0:
                for field_name, field_type, description, example in response_fields:
                    cursor.execute('''
                        INSERT INTO interface_responses (name, response_type, description, example, interface_id)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (field_name, field_type, description, example, interface_id))
            else:
                # 添加默认响应字段
                default_responses = [
                    ('code', 'int', '响应码', '0'),
                    ('message', 'string', '响应消息', 'success'),
                    ('data', 'object', '响应数据', '{}'),
                    ('timestamp', 'string', '时间戳', '2023-01-01 12:00:00')
                ]
                for field_name, field_type, description, example in default_responses:
                    cursor.execute('''
                        INSERT INTO interface_responses (name, response_type, description, example, interface_id)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (field_name, field_type, description, example, interface_id))
        
        # 如果没有匹配到，尝试另一种格式匹配：### 接口名称
# 接口说明
# **POST** /path
        if len(interface_matches) == 0:
            print("尝试第二种匹配格式")
            alt_interface_pattern = r'^(#{2,3})\s+(.+?)\s*(?:\n.*?)?\n\s*(?:\*\*)?\s*(GET|POST|PUT|DELETE|PATCH)\s*(?:\*\*)?\s*[`]?([^\s`]+)[`]?' 
            alt_interface_matches = re.findall(alt_interface_pattern, content, re.DOTALL | re.MULTILINE)
            
            print(f"第二种格式匹配到 {len(alt_interface_matches)} 个接口")
            
            for match in alt_interface_matches:
                header_level, name, method, path = match
                
                # 清理数据
                name = name.strip()
                method = method.upper()
                path = path.strip()
                
                # 如果路径是完整URL，只保留路径部分
                if path.startswith('http'):
                    from urllib.parse import urlparse
                    parsed_url = urlparse(path)
                    path = parsed_url.path
                
                # 打印调试信息
                print(f"解析到接口: {name} {method} {path}")
                
                # 创建接口记录
                cursor.execute('''
                    INSERT INTO interfaces (name, path, method, description, file_id)
                    VALUES (?, ?, ?, ?, ?)
                ''', (name, path, method, '', file_id))
                interface_id = cursor.lastrowid
                parsed_interfaces += 1
                
                # 创建默认Mock配置
                cursor.execute('''
                    INSERT INTO mock_configs (interface_id, enabled, default_count)
                    VALUES (?, ?, ?)
                ''', (interface_id, 1, 10))
                
                # 添加默认请求参数
                default_params = [
                    ('id', 'string', 1, '唯一标识符', '123456'),
                    ('page', 'int', 0, '页码', '1'),
                    ('page_size', 'int', 0, '每页条数', '10')
                ]
                for param_name, param_type, required, description, example in default_params:
                    cursor.execute('''
                        INSERT INTO interface_params (name, param_type, required, description, example, interface_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (param_name, param_type, required, description, example, interface_id))
                parsed_params += len(default_params)
                
                # 添加默认响应字段
                default_responses = [
                    ('code', 'int', '响应码', '0'),
                    ('message', 'string', '响应消息', 'success'),
                    ('data', 'object', '响应数据', '{}'),
                    ('timestamp', 'string', '时间戳', '2023-01-01 12:00:00')
                ]
                for field_name, field_type, description, example in default_responses:
                    cursor.execute('''
                        INSERT INTO interface_responses (name, response_type, description, example, interface_id)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (field_name, field_type, description, example, interface_id))
                parsed_responses += len(default_responses)
        
        # 更新文件记录，保存解析结果统计
        cursor.execute('''
            UPDATE interface_files 
            SET parsed_interfaces = ?, parsed_params = ?, parsed_responses = ? 
            WHERE id = ?
        ''', (parsed_interfaces, parsed_params, parsed_responses, file_id))
        
        conn.commit()
        conn.close()
        
        print(f"文件解析完成: {file_name}, 提取到 {parsed_interfaces} 个接口")
        
    except Exception as e:
        # 清理资源
        if conn:
            conn.rollback()
            conn.close()
        
        print(f"文件解析失败: {file_name}, 错误: {str(e)}")
        
        # 更新文件记录，标记为解析失败
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE interface_files 
            SET parsed = 0, parsed_interfaces = 0, parsed_params = 0, parsed_responses = 0 
            WHERE id = ?
        ''', (file_id,))
        conn.commit()
        conn.close()